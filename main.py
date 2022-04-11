from openpyxl.styles import PatternFill

import Decathlon
import Sportive
import trendyol
import morhipo
import openpyxl
import os
from tkinter import *
from tkinter import ttk


# Delete the current window
def delete(F):
    for element in F.winfo_children():
        element.destroy()

while True:
    # Creating the first Window "GUI"
    window = Tk()
    var = IntVar()
    window.geometry("400x200")
    w = Label(window, text="Choose Url or File", font="Helvetica 16 bold ")
    w.place(x=100, y=50)
    button = Button(window, text="Url", width=25, command=lambda: [var.set(1), delete(window)])
    button.place(x=0, y=100)
    button = Button(window, text="File", width=25, command=lambda: [var.set(2), delete(window)])
    button.place(x=200, y=100)
    button.wait_variable(var)
    Choice = var.get()

    # Creating the second window "GUI"
    var2 = StringVar()
    var3 = StringVar()
    if Choice == 1:
        w = Label(window, text="Enter URl here: ", font="Helvetica 12 bold ")
        w.place(x=110, y=20)
        w = Label(window, text="Enter URl's Code here: ", font="Helvetica 12 bold ")
        w.place(x=30, y=100)
        text_box = Text(window, height=2, width=45)
        text_box.place(x=23, y=50)
        text_box2 = Text(window, height=2, width=20)
        text_box2.place(x=215, y=100)
        button3 = ttk.Button(window, text="Submit", width=25,
                             command=lambda: [var2.set(text_box.get("1.0", "end-1c")),
                                              var3.set(text_box2.get("1.0", "end-1c"))])
    elif Choice == 2:
        w = Label(window, text="Enter File Path here: ", font="Helvetica 16 bold ")
        w.place(x=100, y=50)
        text_box = Text(window, height=2, width=45)
        text_box.place(x=23, y=100)
        button3 = ttk.Button(window, text="Submit", width=25, command=lambda: var2.set(text_box.get("1.0", "end-1c")))

    button3.place(x=100, y=150)
    button3.wait_variable(var2)
    InputCode = var3.get()
    Inputstr = var2.get()
    window.resizable(False, False)
    window.destroy()
    Errors = []
    InputCode = InputCode.replace("\n" , "")
    Inputstr = Inputstr.replace("\n" , "")

    # Creating "Pictures" Path
    path = "Pictures\\"
    try:
        os.mkdir(path)
    except:
        pass

    # Checking the user's choice
    # URL
    if Choice == 1:
        if Inputstr.__contains__("trendyol"):
            trendyol.Trendyol(Inputstr, InputCode)
        elif Inputstr.__contains__("morhipo"):
            morhipo.Morhipo(Inputstr, InputCode)
        elif Inputstr.__contains__("sportive"):
            Sportive.Sportive(Inputstr, InputCode)
        elif Inputstr.__contains__("decathlon"):
            Decathlon.Decathlon(Inputstr, InputCode)
        else:
            print("Unsupported URl")

        # Deleting empty folder
        try:
            os.rmdir("Pictures\\" + InputCode)
            Errors.append(InputCode)
        except:
            pass

        print("---------------------------------------------------\n")

    # Excel File
    elif Choice == 2:

        # Accessing the Excel file
        Inputstr = Inputstr.replace('"', "")
        wb_obj = openpyxl.load_workbook(Inputstr)
        sheet = wb_obj.active
        rows = sheet.max_row
        counter = 0
        for i in range(2, rows + 1):

            # identifying the cells' position
            urlcell = sheet.cell(row=i, column=2)
            Codecell = sheet.cell(row=i, column=1)
            SizeCell = sheet.cell(row=i, column=4)
            PriceCell = sheet.cell(row=i, column=5)
            url = str(urlcell.value)

            if url.__contains__("trendyol"):
                PriceCell.value, SizeCell.value = trendyol.Trendyol(url, str(Codecell.value))

            elif url.__contains__("morhipo"):
                PriceCell.value, SizeCell.value = morhipo.Morhipo(url, str(Codecell.value))

            elif url.__contains__("sportive"):
                PriceCell.value, SizeCell.value = Sportive.Sportive(url, str(Codecell.value))

            elif url.__contains__("decathlon"):
                PriceCell.value, SizeCell.value = Decathlon.Decathlon(url, str(Codecell.value))
            else:
                print("Unsupported URl")
                PriceCell.fill = PatternFill("solid", start_color="00C0C0C0")
                SizeCell.fill = PatternFill("solid", start_color="00C0C0C0")

            counter = counter + 1

            # Color coding
            if PriceCell.value == "Not Available" and SizeCell.value == "Not Available":
                PriceCell.fill = PatternFill("solid", start_color="00FF0000")
                SizeCell.fill = PatternFill("solid", start_color="00FF0000")
            elif SizeCell.value == "Not Available":
                SizeCell.fill = PatternFill("solid", start_color="00FF6600")

            try:
                DirSize = os.listdir("Pictures\\" + Codecell.value + "\\")
                if len(DirSize) == 1:
                    os.remove("Pictures\\" + Codecell.value + "\\" + DirSize[0])
            except:
                pass

            # Deleting empty folders
            try:
                os.rmdir("Pictures\\" + Codecell.value)
                Errors.append(Codecell.value)
            except:
                pass

            print(str(counter) + " / " + str(rows))
            print("---------------------------------------------------\n")

        # Saving file
        wb_obj.save(Inputstr)

    if len(Errors) != 0:
        print("Missing pictures for: ")
        print(Errors)

window.mainloop()
